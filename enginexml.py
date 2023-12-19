import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import os

class Enginexml:
    def __init__(self):

        self.listxmlfile = []
        self.listlayout = []
        self.listlayoutgui = []
        self.listresult = []
        self.listsheetnames = []

        self.filepath = " "
        self.excelfilepath = " "
        self.sheetnames = " "
        self.today = " "
        
        self.username = os.getlogin()
        self.timeknow()
       
        print("Welcome " + self.username)
        print("Date: " + self.today)

# Getting the path of xml conx file 
    def openxmlfile(self, filepath):
        if filepath == "":
            print ("Error!!! xml file (.conx) path is empty. \n FileNotFoundError: [Errno 2] No such file or directory: ' '")
        else:
            self.filepath = filepath
            self.timeknow()
            print("xml file path: \n" + self.filepath + "\nDate: " + self.today)
        return self.filepath

# Getting the path of excel 
    def openexcel(self, excelpath):
        if excelpath == "":
            print("Error!!! EXCEL file path (.xlsm) is empty. \n FileNotFoundError: [Errno 2] No such file or directory: ' '")
        else:
            self.excelfilepath = excelpath
            self.timeknow()
            print("Excel file path: \n" + self.excelfilepath + "\nDate: " + self.today)
            self.sheetnamesexcel()
        return self.excelfilepath

# Read xml conx file to parse information
    def readxmlfile(self, sheetname):
        # change path global variable
        filepathxml = self.filepath
        self.xmlpath = filepathxml
        self.tree = ET.parse(self.xmlpath)
        root = self.tree.getroot()
        for layout in root.findall(".//Layout"):
            namelayout = layout.attrib.get("Name")
            self.listxmlfile.append(namelayout)
            self.listlayout.append(namelayout)
            for path in layout.findall(".//Connection"):
                path = path.attrib.get("Identifier")
                self.listxmlfile.append(path)
        # Delete dublicate items in self.listresult 
        self.listresult = list(dict.fromkeys(self.listxmlfile))
        self.excelsaveopenpyxl(sheetname)

# Write all data in listresult to excel 
    def excelsaveopenpyxl(self, sheetname):
        excelpath = self.excelfilepath
        listresult = self.listresult
        listlayout = self.listlayout
        workbook = load_workbook(excelpath, read_only = False, keep_vba = True)
        sheetname = workbook[sheetname]   
        print(str(sheetname) + " was selected to write all data...")    
        for row, data in enumerate(listresult, start = 4):
            if data in listlayout:
                sheetname.cell(row = row, column = 1).value = data             
                sheetname["A" + str(row)].fill = PatternFill("solid", start_color = "00FFCC00")

            else:
                datasplit=data.split("//")
                # name configuration should be changed...
                sheetname.cell(row = row, column = 1).value = "HIL_XCP_" + str(row-4)
                sheetname.cell(row = row, column = 2).value = "Plant model"
                sheetname.cell(row = row, column = 3).value = datasplit[1]
                if datasplit[1][-4:] == "Out1":
                    sheetname.cell(row = row, column = 4).value = "xaModelSignal"
                elif datasplit[1][-4:] == "/Out":
                    sheetname.cell(row = row, column = 4).value = "xaModelSignal"
                else:
                    sheetname.cell(row = row, column = 4).value = "xaModelValueVariable"
  
        sheetname.cell(row = 3, column = 1).value = "ID"
        sheetname.cell(row = 3, column = 2).value = "MODEL-KEY"
        sheetname.cell(row = 3, column = 3).value = "VARIABLE-PATH"
        sheetname.cell(row = 3, column = 4).value = "TYPE"
        sheetname.cell(row = 3, column = 5).value = "TYPE2"
        sheetname.cell(row = 3, column = 6).value = "DESCRIPTION"
        workbook.save(excelpath)
        print("Excel generated... \n" + self.excelfilepath + "\nDate: " + self.today)

# Get all excel sheetnames to pass GUI
    def sheetnamesexcel(self):
        excelpath = self.excelfilepath
        workbook = load_workbook(excelpath, read_only = False, keep_vba = True)
        sheetnames = workbook.sheetnames
        self.listsheetnames = sheetnames       
        return self.listsheetnames
    
# Get all layout names
    def layoutnames(self):
        filepathxml = self.filepath
        self.xmlpath = filepathxml
        self.tree = ET.parse(self.xmlpath)
        root = self.tree.getroot()
        self.listlayoutgui.clear()
        self.listlayoutgui.append("All Layout")
        for layout in root.findall(".//Layout"):
            namelayout = layout.attrib.get("Name")
            self.listlayoutgui.append(namelayout)
        self.listlayoutgui = list(dict.fromkeys(self.listlayoutgui))
        print("Layout names: \n" + str(self.listlayoutgui))
        return self.listlayoutgui

# Clean all data in the sheet of excel                                                                                                                                
    def cleansheet(self, sheetname):
        excelpath = self.excelfilepath
        workbook = load_workbook(excelpath, read_only = False, keep_vba = True)
        sheetname = workbook[sheetname]
        for row in sheetname["A0:Z5000"]:
            for huc in row:
                huc.value = None
                huc.fill = PatternFill(None, start_color = None) 
        print(str(sheetname) + " has been cleared...")      
        workbook.save(excelpath)      

# Searching data in the excel
    def searchdatainexcel(self, sheetname, layout):
        excelpath = self.excelfilepath
        workbook = load_workbook(excelpath, read_only = False, keep_vba = True)
        sheetname = workbook[sheetname]
        for dataincell in sheetname['A']:
            if dataincell.value == layout:
                print(self.listresult)
                print(self.listresult.index(layout))
            else:
                print("nah")

# Date and hour data
    def timeknow(self):
        self.today = datetime.now().strftime("%d/%m/%Y %H:%M:%S")